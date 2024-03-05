from openpyxl import Workbook, load_workbook

import pandas as pd

import time
# Import the datetime module
from datetime import datetime, date
import math
from copy import copy

"""
We will use the timestamp as package as bar code for packages and store the result to excel file
"""


def timestamp1():
    '''
     # Get the current Unix timestamp as a float
    :return: timestamp
    '''

    now = time.time()
    now = int(now * 10000000)
    # Round the float to the nearest integer
    now_round = round(now)
    return now


def get_quantity(panel, type, Thick):
    """
    :param panel: A datataframe is the excel with the Type and Think
    :param type: order type
    :param Thick: The thinknes of the order
    :return: The quantity for those parameters
    """
    filtered_df = panel.loc[(panel['Type'] == type) & (panel['Thick'] == Thick)]
    quantity = filtered_df.iloc[0]['Quantity']
    return quantity


# define a function to modify the age column based on the gender column


# we have the Thickness of Insulation as constrait Thickness of Insulation

# roof is PR and PT-R
def heigh_(type, qnt, pack_quantity, thick):
    """
    :param type:The type of the package example PR,PT,PW
    :param qnt: PAckage quantity
    :param pack_quantity: The default package quantity for this Thinkess
    :param thick:Get the thickness of the package
    :return: Get height of the package at mm
    """
    insulation = thick
    if type == "PR":
        roof = 20
    else:
        roof = 0

    if (type == "PR" or type == "PT") and (qnt % pack_quantity != 0):
        programming = 20
    else:
        programming = 0

    think_feliz = 100
    # qnt is the quantity of bundle

    return (insulation + roof) * qnt + programming + think_feliz


def width_(type):
    if type == "PR":
        return 1080
    if type == "PT":
        return 1150
    if type == "PW":
        return 1050
    else:
        return 1020


# Get the current date and time
current_date_time = datetime.now()
timestamp = int(current_date_time.timestamp())
print(timestamp)
# Print the result


wb = Workbook()

# grab the active worksheet
ws = wb.active
ws.title = ("Panels Orders")
ws['A1'] = "001panels ordername"
ws['B1'] = current_date_time
ws['A2'] = "BundleNum"
ws['B2'] = "LoT/NuM"
ws['C2'] = "Material"
ws['D2'] = "TEM"
ws['E2'] = "Length"
ws['F2'] = "Τhick "
ws['G2'] = "Width "
ws['H2'] = "Height "
ws['I2'] = "Side "
ws['J2'] = "Color"
wb.save('001panels.xlsx')

order = pd.read_excel("test.xlsx", header=0)
panel = pd.read_excel("PanelsF.xlsx", header=0)
ordern = order.sort_values(['Type', 'Color', 'Side', 'Thick', 'Length'], ascending=[True, True, True, True, False])

orderm = pd.unique(ordern[['Type', 'Color', 'Thick', 'Side', ]].values.ravel('K'))

# order.loc[order['Type'] == 'PR-5T']

# take the unique coloumn or unige type in order
order_type = ordern['Type'].unique().tolist()
order_type_count = order['Type'].nunique()
# for i in order_type:
#   print(i[:2])
#  pan_type=i[:2]
# print(order.loc[order['Type'] == i])


# the coloumn that compare
cols = ["Type", "Color", "Thick", "Side"]

# Compare each row with the previous one and check if all columns are equal
# ordern["Comparison"] = (ordern[cols] == ordern[cols].shift()).all(axis=1).astype(int)
ordern["Comparison"] = (ordern[cols] == ordern[cols].shift(-1)).all(axis=1).astype(int)

# Write DataFrame to Excel
ordern.to_excel("output.xlsx", index=False)

"""
For each row with quantity >0 and if quantity-package -quantity >0 if is true we have the package quantity and save the package barcode and properties for packaage
if the quantity is less than par code and the compare row is 1 we can aggregate the quantity with the next row so we add the quantity to the next package
We store the  package quantity in variable package for these quantities and initialaze when we have package
"""
indx = 3
type_prev = 0
Thick_prev = 0
Length_prev = 0
height = 0
width = 0
Quantity_prev = 0
pack_qnt = 0

grouped_data = {}

for i, row in ordern.iterrows():
        type_ = row["Type"]
        color = row["Color"]
        Thick = row["Thick"]
        Length = row["Length"]
        Quantity = row["Quantity"]
        compare = row["Comparison"]
        Side = row["Side"]
        t = type_[0:2]
        ex = str("pack_qnt=") + "get_quantity(panel,'" + t + "'" + "," + str(Thick) + str(")")
        # print(ex)
        exec(ex)

        indx2 = 0

        while Quantity > 0:
            indx2 = indx2 + 1
            # Check if Quantity is positive
            # Print type and Thick with a comma at the end
            if Quantity - pack_qnt > 0:
                num = indx
                time.sleep(0.0000001)
                ws['A' + str(num)] = str(timestamp1())
                ws['B' + str(num)] = indx2
                ws['C' + str(num)] = type_
                q = pack_qnt - Quantity_prev
                ws['D' + str(num)] = q
                ws['E' + str(num)] = Length
                ws['F' + str(num)] = Thick
                ws['I' + str(num)] = Side
                ws['J' + str(num)] = color
                # calculate the height
                ex2 = str("height=") + "heigh_('" + t + "'" + "," + str(q) + "," + str(pack_qnt) + "," + str(Thick) + str(
                    ")")
                # print(ex2)
                exec(ex2)
                ws['H' + str(num)] = height

                # calculate width
                ex3 = str("width=") + "width_('" + t + "'" + str(")")
                # print(ex3)
                exec(ex3)

                ws['G' + str(num)] = width

                Quantity = Quantity - pack_qnt  # Decrease Quantity by 1
                pakage = 0

                indx = indx + 1
                Quantity_prev = 0
            else:
                num = indx
                # if we canot
                if compare == 0:
                    time.sleep(0.0000001)
                    ws['A' + str(num)] = str(timestamp1())
                    ws['B' + str(num)] = indx2
                    ws['C' + str(num)] = type_
                    ws['D' + str(num)] = Quantity
                    ws['E' + str(num)] = Length
                    ws['F' + str(num)] = Thick
                    ws['I' + str(num)] = Side
                    ws['J' + str(num)] = color

                    # calculate the height
                    ex2 = str("height=") + "heigh_('" + t + "'" + "," + str(Quantity) + "," + str(pack_qnt) + "," + str(
                        Thick) + str(")")
                    # print(ex2)
                    exec(ex2)
                    ws['H' + str(num)] = height

                    # calculate width
                    ex3 = str("width=") + "width_('" + t + "'" + str(")")
                    # print(ex3)
                    exec(ex3)

                    ws['G' + str(num)] = width

                    pakage = 0
                    Quantity = 0
                    indx = indx + 1
                    Quantity_prev = 0
                else:
                    if (pakage + Quantity) < pack_qnt:
                        type_prev = type_
                        Thick_prev = Thick
                        Length_prev = Length
                        Quantity_prev = Quantity
                        pakage = pakage + Quantity
                        time.sleep(0.0000001)
                        ws['A' + str(num)] = "same_package"
                        ws['B' + str(num)] = indx2
                        ws['C' + str(num)] = type_
                        ws['D' + str(num)] = Quantity
                        ws['E' + str(num)] = Length
                        ws['F' + str(num)] = Thick
                        ws['I' + str(num)] = Side
                        ws['J' + str(num)] = color
                        # calculate the height
                        ex2 = str("height=") + "heigh_('" + t + "'" + "," + str(Quantity) + "," + str(pack_qnt) + "," + str(
                            Thick) + str(")")
                        # print(ex2)
                        exec(ex2)
                        ws['H' + str(num)] = height

                        # calculate width
                        ex3 = str("width=") + "width_('" + t + "'" + str(")")
                        # print(ex3)
                        exec(ex3)

                        ws['G' + str(num)] = width

                        indx = indx + 1
                        Quantity = 0
                    else:
                        time.sleep(0.0000001)
                        ws['A' + str(num)] = str(timestamp1())
                        ws['B' + str(num)] = indx2
                        ws['C' + str(num)] = type_
                        ws['D' + str(num)] = Quantity
                        ws['E' + str(num)] = Length
                        ws['F' + str(num)] = Thick
                        ws['I' + str(num)] = Side
                        ws['J' + str(num)] = color

                        # calculate the height
                        ex2 = str("height=") + "heigh_('" + t + "'" + "," + str(Quantity) + "," + str(pack_qnt) + "," + str(
                            Thick) + str(")")
                        # print(ex2)
                        exec(ex2)
                        ws['H' + str(num)] = height

                        # calculate width
                        ex3 = str("width=") + "width_('" + t + "'" + str(")")
                        # print(ex3)
                        exec(ex3)

                        ws['G' + str(num)] = width

                        pakage = 0
                        Quantity = 0
                        indx = indx + 1
                        pakage = 0

        wb.save('001panels.xlsx')
last_col = ws.max_column

"""
    Now we have to write the excel the packages file.
    We split the area of excel into ranges that can be store files
    All the ranges are 30. WE have 6 list of sections , The section1, section2, section3 and section4 and 5 and 6
    we have to find how many ranges is going to use each section with  the len(list)/2 and all divide by 12 because we have 
    12 emplty fields each cell ranges. Example if the section len 82 divide by 2 we have tuples, is 41 divided be 12 is 3.416 roudUp is 4
    for this setion we are going to use 4 cell ranges
"""

zb = load_workbook('Package.xlsx')
zs = zb.active
zs = zb["ΔΕΜΑΤΟΠΟΙΗΣΗ"]
cell_range1 = zs['C9':'D20']
cell_range2 = zs['H9':'I20']
cell_range3 = zs['M9':'N20']
cell_range4 = zs['R9':'S20']
cell_range5 = zs['W9':'X20']
cell_range6 = zs['C33':'D44']
cell_range7 = zs['H33':'I44']
cell_range8 = zs['M33':'N44']
cell_range9 = zs['R33':'S44']
cell_range10 = zs['W33':'X44']
cell_range11 = zs['C62':'D73']
cell_range12 = zs['H62':'I73']
cell_range13 = zs['M62':'N73']
cell_range14 = zs['R62':'S73']
cell_range15 = zs['W62':'X73']
cell_range16 = zs['C86':'D97']
cell_range17 = zs['H86':'I97']
cell_range18 = zs['M86':'N97']
cell_range19 = zs['R86':'S97']
cell_range20 = zs['W86':'X97']
cell_range21 = zs['C115':'D20']
cell_range22 = zs['H115':'D20']
cell_range23 = zs['M115':'D20']
cell_range24 = zs['R115':'D20']
cell_range25 = zs['W115':'D20']
cell_range26 = zs['C139':'D20']
cell_range27 = zs['H139':'D20']
cell_range28 = zs['M139':'D20']
cell_range29 = zs['R139':'D20']
cell_range30 = zs['W139':'D20']

header1 = "E7"
header2 = 'J7'
header3 = 'O7'
header4 = 'E7'
header5 = 'T7'
header6 = 'Y7'
header7 = 'E31'
header8 = 'J31'
header9 = 'O31'
header10 = 'T31'
header11 = 'Y31'
header12 = 'E60'
header13 = 'J60'
header14 = 'O60'
header15 = 'T60'
header16 = 'Y60'

# Get the cell range

cell_ran = ws['A3':'I' + str(ws.max_row)]

# Create an empty list to store the values if length and quantity to fill in the package excel
section1 = []
section2 = []
section3 = []
section4 = []
section5 = []
section6 = []

cell_list = []  # create an empty list
counter = 0
previous_C = 0
previous_J = 0

"""
We are going to read from excel and group the data together where thay are at the same group.
First we grouping the data we read to packages
"""

# Dictionary to store grouped data
grouped_data = {}
# Helper variables to track grouping
previous_E = []
previous_D = []
previous_H = []
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=10):
    cell_A = row[0]  # get the cell in column package id
    cell_C = row[2]  # get the cell in column package type
    cell_I = row[8]  # get the cell in column I side
    cell_E = row[4]  # get the cell in column E lenght
    cell_F = row[5]  # get the cell in column F Thik
    cell_G = row[6]  # get the cell in column G Width
    cell_D = row[3]  # get the cell in column D quantity temaxia
    cell_H = row[7]  # get the cell in column H height
    cell_J = row[9]  # get the cell in column H color
    type_m=str(cell_C.value)+str(cell_J.value)
    if cell_A.value != "same_package":
        grouped_data[cell_A.value]=[]
        grouped_data[cell_A.value].append(cell_E.value)
        grouped_data[cell_A.value].append(cell_D.value)
        if len(previous_D)!=0:
            for i in range(0,len(previous_E)):
                grouped_data[cell_A.value].append(previous_E.pop())
                grouped_data[cell_A.value].append(previous_D.pop())
            previous_E = []
            previous_D = []

    if cell_A.value == "same_package":
        previous_E.append(cell_E.value)
        previous_D.append(cell_D.value)
    data = grouped_data
    df = pd.DataFrame.from_dict(data, orient='index')
    df = df.reset_index()
    df.columns = ['ID'] + [f'Value{i}' for i in range(1, len(df.columns))]
    df.to_excel('output.xlsx', index=False)





    zs[header1] = cell_I.value

    if (cell_C.value != previous_C) or (cell_J.value != previous_J):
        counter += 1



        # crreate new sheet and copy the previous sheet's data
        new_ws = zb.copy_worksheet(zs)
        new_ws.title = cell_C.value
        zb.active = new_ws


        zs[header1] = cell_I.value

        previous_C = cell_C.value
        previous_J = cell_J.value

    zb.save('Package.xlsx')

zb.save('Package.xlsx')
